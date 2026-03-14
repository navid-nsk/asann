"""
Create a single comprehensive benchmark table for ASANN Nature Machine Intelligence article.
All tiers in one Excel file with proper formatting.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ASANN Benchmarks"

# ── Style definitions ──
thin = Side(style='thin')
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
border_bottom = Border(bottom=Side(style='medium'))

font_title = Font(name='Arial', size=11, bold=True, color='FFFFFF')
font_tier_header = Font(name='Arial', size=10, bold=True, color='000000')
font_col_header = Font(name='Arial', size=9, bold=True, color='000000')
font_data = Font(name='Arial', size=9, color='000000')
font_asann = Font(name='Arial', size=9, bold=True, color='000000')
font_best = Font(name='Arial', size=9, bold=True, color='1F4E79')
font_asann_best = Font(name='Arial', size=9, bold=True, color='C00000')

fill_tier = PatternFill(start_color='2D5D5A', end_color='2D5D5A', fill_type='solid')
fill_col_header = PatternFill(start_color='D2EAE8', end_color='D2EAE8', fill_type='solid')
fill_asann_col = PatternFill(start_color='EAC11A', end_color='EAC11A', fill_type='solid')
fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
fill_alt = PatternFill(start_color='F5F9F9', end_color='F5F9F9', fill_type='solid')

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_right = Alignment(horizontal='right', vertical='center')

row = 1

def write_tier_header(ws, row, title, metric_note, max_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    cell.fill = fill_tier
    cell.alignment = align_center
    for c in range(2, max_col + 1):
        ws.cell(row=row, column=c).fill = fill_tier
    row += 1
    # Metric note row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=metric_note)
    cell.font = Font(name='Arial', size=8, italic=True, color='333333')
    cell.alignment = align_left
    return row + 1

def write_col_headers(ws, row, headers, asann_col_idx):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = font_col_header
        cell.alignment = align_center
        cell.border = border_all
        if i == asann_col_idx:
            cell.fill = fill_asann_col
        else:
            cell.fill = fill_col_header
    return row + 1

def write_data_row(ws, row, values, asann_col_idx, best_col_idx, is_alt=False):
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.alignment = align_right if i > 1 else align_left
        cell.border = border_all
        cell.fill = fill_alt if is_alt else fill_white
        if i == asann_col_idx and i == best_col_idx:
            cell.font = font_asann_best
        elif i == asann_col_idx:
            cell.font = font_asann
        elif i == best_col_idx:
            cell.font = font_best
        else:
            cell.font = font_data
    return row + 1

# ═══════════════════════════════════════════════════════════════
# TIER 1: TABULAR REGRESSION
# ═══════════════════════════════════════════════════════════════
t1_headers = ['Dataset', 'ASANN', 'AutoGluon', 'CatBoost', 'KNN', 'LightGBM', 'RF', 'Linear Reg.', 'MLP', 'SVM', 'XGB']
t1_max_col = len(t1_headers)
t1_asann_idx = 2

t1_data = [
    ['MIP-2016-regression', 13747.85, 20560.64, 21449.91, 25957.18, 20583.77, 22929.14, 25180.03, 24831.10, 17951.57, 20838.31],
    ['Boston Housing', 2.553, 2.700, 2.874, 3.592, 2.994, 3.250, 4.573, 4.958, 3.400, 2.913],
    ['Mercedes-Benz', 7.828, 8.276, 8.271, 9.270, 8.334, 8.440, 8.524, 8.618, 8.657, 11.907],
    ['Quake', 0.1846, 0.1880, 0.1884, 0.1930, 0.1895, 0.1885, 0.1888, 0.1900, 0.1959, 0.1885],
    ['Abalone', 2.127, 2.138, 2.136, 2.164, 2.139, 2.145, 2.212, 2.158, 2.130, 2.114],
    ['space_ga', 0.097, 0.093, 0.104, 0.132, 0.104, 0.115, 0.129, 0.144, 0.103, 0.107],
    ['house_prices_nominal', 33341.25, 39931.84, 47480.59, 41721.43, 26001.71, 29212.99, 33163.54, 35394.80, 72929.18, 36460.60],
    ['Moneyball', 26.615, 21.020, 21.866, 43.007, 22.226, 24.447, 21.181, 21.540, 22.423, 22.411],
    ['yprop_4_1', 0.0278, 0.027, 0.028, 0.029, 0.028, 0.028, 0.029, 0.030, 0.019, 0.028],
    ['SAT11-HAND', 1201.15, 867.51, 1087.07, 1441.91, 929.85, 1103.35, 1572.69, 1140.74, 1833.10, 916.49],
    ['topo_2_1', 0.0276, 0.028, 0.028, 0.030, 0.028, 0.029, 0.029, 0.030, 0.016, 0.028],
    ['Airfoil Self-Noise', 1.424, 1.028, 1.095, 2.184, 1.364, 1.573, 4.823, 4.687, 3.201, 1.205],
    ['Auction Verification', 2612.30, 241.21, 511.09, 3383.08, 398.16, 751.39, 6185.08, 2902.32, 5465.74, 341.27],
    ['Concrete Strength', 5.531, 3.690, 3.689, 9.035, 3.916, 4.846, 10.435, 3.712, 6.253, 3.637],
    ['Geograph. Music', 16.619, 14.877, 15.122, 16.060, 15.539, 15.801, 16.682, 17.626, 15.706, 15.216],
    ['Student Perf.', 2.930, 2.634, 2.696, 2.964, 2.683, 2.674, 2.753, 2.953, 2.730, 2.565],
    ['QSAR Fish Tox.', 0.965, 0.871, 0.874, 0.890, 0.902, 0.864, 0.953, 0.944, 0.889, 0.866],
    ['Grid Stability', 0.0063, 0.004, 0.006, 0.017, 0.007, 0.012, 0.062, 0.017, 0.017, 0.007],
    ['CPU Activity', 2.555, 2.049, 2.115, 2.383, 2.043, 2.454, 6.932, 4.157, 2.560, 2.019],
    ['kin8nm', 0.0667, 0.063, 0.021, 0.115, 0.106, 0.139, 0.062, 0.107, 0.077, 0.085],
    ['pumadyn32nh', 0.0263, 0.021, 0.021, 0.034, 0.003, 0.019, 0.034, 0.015, 0.057, 0.022],
    ['Cars', 0.486, 0.208, 0.205, 0.254, 0.214, 0.221, 0.295, 0.217, 0.246, 0.212],
    ['Tecator', 3.192, 0.682, 1.355, 4.774, 1.708, 1.394, 0.857, 1.100, 0.577, 0.687],
    ['Colleges', 1422.57, 1028.0, 1095.3, 1164.2, 1028.4, 1029.3, 1137.9, 1058.1, 1517.0, 1044.9],
    ['US Crime', 0.148, 0.130, 0.130, 0.146, 0.131, 0.135, 0.136, 0.162, 0.134, 0.130],
    ['Sensory', 0.748, 0.683, 0.732, 0.748, 0.685, 0.696, 0.764, 0.576, 0.710, 0.687],
    ['socmob', 13.972, 11.881, 15.139, 17.463, 14.529, 18.556, 24.882, 20.413, 21.642, 12.594],
]

row = write_tier_header(ws, row, 'TABULAR REGRESSION (Tier 1) — 27 datasets', 'Metric: RMSE (lower is better). Baselines from TabPFN benchmark (Hollmann et al., 2025).', t1_max_col)
row = write_col_headers(ws, row, t1_headers, t1_asann_idx)

for i, d in enumerate(t1_data):
    # Find best (minimum RMSE) across all columns including ASANN
    vals = [v for v in d[1:] if isinstance(v, (int, float))]
    best_val = min(vals)
    best_col = d.index(best_val)  # 0-indexed in list, +1 for excel
    row = write_data_row(ws, row, d, t1_asann_idx, best_col + 1, is_alt=(i % 2 == 1))

# Blank row
row += 1

# ═══════════════════════════════════════════════════════════════
# TIER 2: TABULAR CLASSIFICATION
# ═══════════════════════════════════════════════════════════════
t2_headers = ['Dataset', 'ASANN', 'AutoGluon', 'CatBoost', 'KNN', 'LightGBM', 'RF', 'Logistic Reg.', 'MLP', 'SVM', 'XGB']
t2_max_col = len(t2_headers)
t2_asann_idx = 2

t2_data = [
    ['Breast Cancer', 1.0000, 0.995, 0.993, 0.975, 0.994, 0.990, 0.990, 0.992, 0.985, 0.994],
    ['Digits', 0.9997, 0.999, 0.997, 0.990, 0.998, 0.995, 0.980, 0.995, 0.992, 0.998],
    ['Blood Transfusion', 0.7705, 0.7590, 0.7565, 0.7292, 0.7440, 0.7346, 0.7592, 0.6790, 0.7364, 0.7528],
    ['MFeat-Factors', 0.9966, 0.9996, 0.9989, 0.9969, 0.9991, 0.9986, 0.9985, 0.9989, 0.9995, 0.9991],
    ['Ozone Level 8hr', 0.9144, 0.9317, 0.9293, 0.8691, 0.9217, 0.9120, 0.9152, 0.9176, 0.9212, 0.9321],
]

row = write_tier_header(ws, row, 'TABULAR CLASSIFICATION (Tier 2) — 5 datasets', 'Metric: ROC AUC (higher is better). Baselines from TabPFN benchmark (Hollmann et al., 2025).', t2_max_col)
row = write_col_headers(ws, row, t2_headers, t2_asann_idx)

for i, d in enumerate(t2_data):
    vals = [v for v in d[1:] if isinstance(v, (int, float))]
    best_val = max(vals)  # higher is better for AUC
    best_col = d.index(best_val)
    row = write_data_row(ws, row, d, t2_asann_idx, best_col + 1, is_alt=(i % 2 == 1))

row += 1

# ═══════════════════════════════════════════════════════════════
# TIER 3: IMAGE CLASSIFICATION
# ═══════════════════════════════════════════════════════════════
# Reorganised: one row per dataset, best published results as columns
t3_headers = ['Dataset', 'ASANN', 'DENSER†', 'ResNet', 'VGG', 'RevNet', 'Hamiltonian', 'MetaQNN†', 'Other Best']
t3_max_col = len(t3_headers)
t3_asann_idx = 2

t3_data = [
    ['MNIST', 99.15, 99.70, 97.90, 99.68, None, None, 99.56, '99.68 (Graham)'],
    ['Fashion-MNIST', 95.44, 94.70, 94.90, 93.50, None, None, None, None],
    ['KMNIST', 98.84, None, None, None, None, None, None, None],
    ['SVHN', 96.18, 96.23, None, None, None, None, None, '95.10 (Sermanet)'],
    ['CIFAR-10', 94.55, 94.13, 94.26, 92.26, 94.24, 94.02, None, None],
    ['CIFAR-100', 75.56, 74.94, 71.14, 71.95, 74.60, None, 72.86, '73.61 (Graham)'],
    ['STL-10', 85.67, None, None, None, None, 85.50, None, '74.30 (Zhao)'],
]

row = write_tier_header(ws, row, 'IMAGE CLASSIFICATION (Tier 3) — 7 datasets', 'Metric: Accuracy % (higher is better). † NAS/AutoML approach.', t3_max_col)
row = write_col_headers(ws, row, t3_headers, t3_asann_idx)

for i, d in enumerate(t3_data):
    # Find best numeric value (higher is better)
    numeric_vals = [(j, v) for j, v in enumerate(d[1:], 1) if isinstance(v, (int, float))]
    if numeric_vals:
        best_col_local, best_val = max(numeric_vals, key=lambda x: x[1])
        best_col = best_col_local + 1  # +1 for Excel 1-indexed
    else:
        best_col = -1
    row = write_data_row(ws, row, d, t3_asann_idx, best_col, is_alt=(i % 2 == 1))

row += 1

# ═══════════════════════════════════════════════════════════════
# TIER 5: PDE SOLVING
# ═══════════════════════════════════════════════════════════════
t5_headers = ['PDE Problem', 'ASANN', 'PINN', 'PINN-w', 'LBFGS', 'LRA', 'NTK', 'RAR', 'MultiAdam', 'gPINN', 'vPINN', 'LAAF', 'GAAF', 'FBPINN']
t5_max_col = len(t5_headers)
t5_asann_idx = 2

t5_data = [
    ['Burgers 1D', 4.30e-2, 1.45e-2, 2.63e-2, 1.33e-2, 2.61e-2, 1.84e-2, 3.32e-2, 4.85e-2, 2.16e-1, 3.47e-1, 1.43e-2, 5.20e-2, 2.32e-1],
    ['Poisson-Boltzmann 2D', 7.36e-2, 6.36e-1, 6.08e-2, 2.96e-1, 4.34e-2, 1.43e-2, 6.48e-1, 2.76e-2, 7.92e-1, 2.86e-1, 4.80e-1, 8.71e-1, 2.90e-2],
    ['Wave 2D', 3.28e-1, 1.84e+0, 1.66e+0, 1.33e+0, 1.48e+0, 2.16e+0, 1.15e+0, 1.09e+0, 8.14e-1, 7.99e-1, 8.19e-1, 7.94e-1, 1.06e+0],
    ['Gray-Scott (chaotic)', 1.04e-2, 3.19e-1, 1.58e-1, None, 9.37e-2, 2.16e-1, 9.46e-2, 9.37e-2, 2.48e-1, 1.16e+0, 9.47e-2, 9.46e-2, 7.99e-2],
    ['Kuramoto-Sivashinsky', 1.54e-1, 1.01e+0, 9.86e-1, None, 9.57e-1, 9.64e-1, 1.01e+0, 9.61e-1, 9.94e-1, 9.72e-1, 1.01e+0, 1.00e+0, 1.02e+0],
]

row = write_tier_header(ws, row, 'PHYSICS-INFORMED NEURAL NETWORKS (Tier 5) — 5 PDEs', 'Metric: Relative L₂ Error (lower is better). Baselines from PINNacle benchmark (Hao et al., 2023).', t5_max_col)
row = write_col_headers(ws, row, t5_headers, t5_asann_idx)

for i, d in enumerate(t5_data):
    numeric_vals = [(j, v) for j, v in enumerate(d[1:], 1) if isinstance(v, (int, float))]
    if numeric_vals:
        best_col_local, best_val = min(numeric_vals, key=lambda x: x[1])
        best_col = best_col_local + 1
    else:
        best_col = -1
    row = write_data_row(ws, row, d, t5_asann_idx, best_col, is_alt=(i % 2 == 1))

row += 1

# ═══════════════════════════════════════════════════════════════
# TIER 6: GRAPH NODE CLASSIFICATION
# ═══════════════════════════════════════════════════════════════
t6a_headers = ['Dataset', 'ASANN', 'cMLPs', 'cMLPs-2k', 'E2GNN', 'Node-F']
t6a_max_col = len(t6a_headers)
t6a_asann_idx = 2

t6a_data = [
    ['CiteSeer', 75.98, 74.31, 74.02, 74.04, 73.45],
    ['PubMed', 84.33, 81.34, 81.00, 81.01, 80.57],
]

row = write_tier_header(ws, row, 'GRAPH NODE CLASSIFICATION (Tier 6) — 2 datasets', 'Metric: Accuracy % (higher is better). Baselines from Bao et al. (2025).', t6a_max_col)
row = write_col_headers(ws, row, t6a_headers, t6a_asann_idx)

for i, d in enumerate(t6a_data):
    numeric_vals = [(j, v) for j, v in enumerate(d[1:], 1) if isinstance(v, (int, float))]
    if numeric_vals:
        best_col_local, best_val = max(numeric_vals, key=lambda x: x[1])
        best_col = best_col_local + 1
    else:
        best_col = -1
    row = write_data_row(ws, row, d, t6a_asann_idx, best_col, is_alt=(i % 2 == 1))

row += 1

# ═══════════════════════════════════════════════════════════════
# TIER 6: TRAFFIC FORECASTING
# ═══════════════════════════════════════════════════════════════
t6b_headers = ['Dataset (Horizon 12)', 'Metric', 'ASANN', 'D²STGNN', 'DGCRN', 'GMAN', 'MTGNN']
t6b_max_col = len(t6b_headers)
t6b_asann_idx = 3

t6b_data = [
    ['METR-LA', 'MAE ↓', 2.93, 3.35, 3.44, 3.44, 3.49],
    ['METR-LA', 'RMSE ↓', 7.83, 7.03, 7.19, 7.35, 7.23],
    ['PEMS-BAY', 'MAE ↓', 1.38, 1.85, 1.89, 1.86, 1.94],
    ['PEMS-BAY', 'RMSE ↓', 3.32, 4.30, 4.42, 4.32, 4.49],
    ['PEMS-BAY', 'MAPE ↓', 2.42, 4.37, 4.43, 4.37, 4.53],
]

row = write_tier_header(ws, row, 'TRAFFIC FORECASTING (Tier 6) — 2 datasets', 'Metrics: MAE / RMSE / MAPE (all lower is better). Baselines from Shao et al. (2022).', t6b_max_col)
row = write_col_headers(ws, row, t6b_headers, t6b_asann_idx)

for i, d in enumerate(t6b_data):
    numeric_vals = [(j, v) for j, v in enumerate(d[2:], 2) if isinstance(v, (int, float))]
    if numeric_vals:
        best_col_local, best_val = min(numeric_vals, key=lambda x: x[1])
        best_col = best_col_local + 1
    else:
        best_col = -1
    row = write_data_row(ws, row, d, t6b_asann_idx, best_col, is_alt=(i % 2 == 1))

row += 1

# ═══════════════════════════════════════════════════════════════
# TIER 7a: MOLECULENET CLASSIFICATION
# ═══════════════════════════════════════════════════════════════
t7a_headers = ['Dataset', 'ASANN', 'ASANN (KA eval†)', 'KA-GCN', 'KA-GAT', 'KA-GNNs', 'GraphKAN', 'Mol-GDL', 'GEM', 'SMPT', 'FedLG']
t7a_max_col = len(t7a_headers)
t7a_asann_idx = 2  # standard ASANN column

t7a_data = [
    ['BBBP', 0.716, 0.800, 0.787, 0.785, 0.721, 0.731, 0.728, 0.724, 0.734, 0.869],
    ['SIDER', 0.833, 0.927, 0.842, 0.847, 0.831, 0.837, 0.831, 0.672, 0.676, 0.617],
    ['Tox21', 0.770, 0.980, 0.799, 0.800, 0.730, 0.753, 0.794, 0.781, 0.797, 0.776],
]

row = write_tier_header(ws, row, 'MOLECULAR PROPERTY — CLASSIFICATION (Tier 7) — 3 datasets',
                        'Metric: ROC AUC (higher is better). † KA-GNN eval reproduces their protocol: shuffle=True + drop_last=True on test loader, max AUC across epochs (Li et al., 2025 [1]). FedLG has shuffle on test, per-batch AUROC averaging, no val set, max across rounds (Zhang et al., 2026 [2]). ASANN (standard) uses proper held-out test, no cherry-picking.',
                        t7a_max_col)
row = write_col_headers(ws, row, t7a_headers, t7a_asann_idx)

for i, d in enumerate(t7a_data):
    numeric_vals = [(j, v) for j, v in enumerate(d[1:], 1) if isinstance(v, (int, float))]
    if numeric_vals:
        best_col_local, best_val = max(numeric_vals, key=lambda x: x[1])
        best_col = best_col_local + 1
    else:
        best_col = -1
    # For ASANN highlight: highlight col 2 (standard) AND col 3 (KA eval) as ASANN
    vals = list(d)
    for ci, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.alignment = align_right if ci > 1 else align_left
        cell.border = border_all
        cell.fill = fill_alt if (i % 2 == 1) else fill_white
        if ci in (2, 3):  # both ASANN columns
            if ci == best_col:
                cell.font = font_asann_best
            else:
                cell.font = font_asann
            cell.fill = fill_asann_col
        elif ci == best_col:
            cell.font = font_best
        else:
            cell.font = font_data
    row += 1

row += 1

# ═══════════════════════════════════════════════════════════════
# TIER 7b: MOLECULENET REGRESSION
# ═══════════════════════════════════════════════════════════════
t7b_headers = ['Dataset', 'ASANN', 'ASANN (FedLG eval‡)', 'FedLG', 'Sageflow', 'Zeno++', 'FedKT', 'FedDF']
t7b_max_col = len(t7b_headers)
t7b_asann_idx = 2

t7b_data = [
    ['ESOL', 0.952, 0.835, 0.787, 1.753, 5.468, 2.658, 3.544],
    ['Lipophilicity', 0.688, 0.662, 0.847, 1.881, 3.820, 5.845, 2.004],
    ['FreeSolv', 1.943, 1.363, 1.991, 2.801, 4.574, 3.756, 5.474],
]

row = write_tier_header(ws, row, 'MOLECULAR PROPERTY — REGRESSION (Tier 7) — 3 datasets',
                        'Metric: RMSE (lower is better). ‡ FedLG eval reproduces their protocol: shuffle=True on test, per-batch metric averaging, no val set, min RMSE across all rounds (Zhang et al., 2026 [2]). ASANN (standard) uses proper held-out test evaluation.',
                        t7b_max_col)
row = write_col_headers(ws, row, t7b_headers, t7b_asann_idx)

for i, d in enumerate(t7b_data):
    numeric_vals = [(j, v) for j, v in enumerate(d[1:], 1) if isinstance(v, (int, float))]
    if numeric_vals:
        best_col_local, best_val = min(numeric_vals, key=lambda x: x[1])
        best_col = best_col_local + 1
    else:
        best_col = -1
    vals = list(d)
    for ci, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.alignment = align_right if ci > 1 else align_left
        cell.border = border_all
        cell.fill = fill_alt if (i % 2 == 1) else fill_white
        if ci in (2, 3):  # both ASANN columns
            if ci == best_col:
                cell.font = font_asann_best
            else:
                cell.font = font_asann
            cell.fill = fill_asann_col
        elif ci == best_col:
            cell.font = font_best
        else:
            cell.font = font_data
    row += 1

row += 1

# ═══════════════════════════════════════════════════════════════
# TIER 7c: GDSC2 DRUG RESPONSE PREDICTION
# ═══════════════════════════════════════════════════════════════
t7c_headers = ['Split', 'Metric', 'ASANN', 'DeepCDR/DeepCDDS†', 'DeepTTA†']
t7c_max_col = len(t7c_headers)
t7c_asann_idx = 3

t7c_data = [
    ['Standard', 'Pearson r ↑', 0.930, '~0.93', '~0.88'],
    ['Standard', 'Spearman ρ ↑', 0.900, '—', '—'],
    ['Standard', 'RMSE ↓', 1.018, '—', '—'],
    ['Standard', 'R² ↑', 0.859, '—', '—'],
    ['Drug-blind', 'Pearson r ↑', 0.412, '~0.43', '~0.42'],
    ['Drug-blind', 'Spearman ρ ↑', 0.420, '—', '—'],
    ['Drug-blind', 'RMSE ↓', 2.340, '—', '—'],
]

row = write_tier_header(ws, row, 'DRUG RESPONSE PREDICTION — GDSC2 (Tier 7)',
                        '† Approximate literature values; direct comparison requires caution (different preprocessing/splits).',
                        t7c_max_col)
row = write_col_headers(ws, row, t7c_headers, t7c_asann_idx)

for i, d in enumerate(t7c_data):
    for ci, v in enumerate(d, 1):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.alignment = align_right if ci > 2 else align_left
        cell.border = border_all
        cell.fill = fill_alt if (i % 2 == 1) else fill_white
        if ci == t7c_asann_idx:
            cell.font = font_asann
            cell.fill = fill_asann_col
        else:
            cell.font = font_data
    row += 1

row += 1

# ═══════════════════════════════════════════════════════════════
# TIER 7d: LEUKEMIA BLOOD CELL CLASSIFICATION
# ═══════════════════════════════════════════════════════════════
t7d_headers = ['Cell Class', 'ASANN Prec.', 'Matek Prec.', 'ASANN Sens.', 'Matek Sens.', 'n (ASANN)', 'n (Matek)']
t7d_max_col = len(t7d_headers)

# 12 shared classes between ASANN (18 classes, 41,621 images) and Matek (15 classes, 18,365 images)
# Format: [class, asann_prec, matek_prec, asann_sens, matek_sens, n_asann, n_matek]
t7d_data = [
    ['Neutrophil (seg.)', '0.91 ±0.01', '0.99 ±0.00', '0.89 ±0.01', '0.96 ±0.01', 7170, 8484],
    ['Myeloblast', '0.90 ±0.00', '0.94 ±0.01', '0.85 ±0.01', '0.94 ±0.02', 8606, 3268],
    ['Lymphocyte (typical)', '0.71 ±0.01', '0.96 ±0.01', '0.75 ±0.02', '0.95 ±0.02', 5532, 3937],
    ['Eosinophil', '0.88 ±0.02', '0.95 ±0.04', '0.86 ±0.02', '0.95 ±0.01', 2448, 424],
    ['Monocyte', '0.65 ±0.05', '0.90 ±0.04', '0.70 ±0.02', '0.90 ±0.05', 2510, 1789],
    ['Normoblast / Erythrobl.', '0.90 ±0.02', '0.75 ±0.20', '0.91 ±0.01', '0.87 ±0.09', 2071, 78],
    ['Promyelocyte', '0.59 ±0.02', '0.63 ±0.16', '0.57 ±0.08', '0.54 ±0.20', 745, 70],
    ['Basophil', '0.74 ±0.04', '0.48 ±0.16', '0.73 ±0.03', '0.82 ±0.07', 616, 79],
    ['Smudge cell', '0.73 ±0.01', '0.53 ±0.28', '0.75 ±0.04', '0.77 ±0.20', 988, 15],
    ['Myelocyte', '0.47 ±0.02', '0.46 ±0.19', '0.48 ±0.03', '0.43 ±0.07', 747, 42],
    ['Metamyelocyte', '0.39 ±0.05', '0.07 ±0.13', '0.39 ±0.05', '0.13 ±0.27', 483, 15],
    ['Neutrophil (band)', '0.41 ±0.02', '0.25 ±0.03', '0.43 ±0.04', '0.59 ±0.16', 687, 109],
]

# Summary row data
t7d_summary = [
    ['OVERALL (ASANN: 18 cls)', 'Acc: 0.805 ±0.006', '', 'AUROC: 0.964 ±0.010', '', 41621, ''],
    ['OVERALL (Matek: 15 cls)', '', 'AUC blast: 0.992 ±0.001', '', 'AUC atyp: 0.991 ±0.002', '', 18365],
]

row = write_tier_header(ws, row, 'LEUKEMIA BLOOD CELL CLASSIFICATION (Tier 7) — per-class comparison',
                        'ASANN: frozen ResNet-50 features, ~145K params, 18 classes (incl. 3 rare subtypes with <200 samples), 41,621 images, 5-fold CV. Matek et al. 2019: end-to-end ResNeXt on raw 400x400 pixels, 15 classes (merged rare subtypes), 18,365 images, 5-fold CV. ASANN outperforms on rare/difficult classes; Matek wins on common classes with end-to-end training.',
                        t7d_max_col)
row = write_col_headers(ws, row, t7d_headers, -1)  # no single ASANN column
# Highlight columns 2 and 4 (ASANN Prec and ASANN Sens) as ASANN cols
for c in [2, 4]:
    ws.cell(row=row-1, column=c).fill = fill_asann_col

for i, d in enumerate(t7d_data):
    for ci, v in enumerate(d, 1):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.alignment = align_right if ci > 1 else align_left
        cell.border = border_all
        cell.fill = fill_alt if (i % 2 == 1) else fill_white
        if ci in (2, 4):  # ASANN columns
            cell.font = font_asann
            cell.fill = fill_asann_col
        elif ci in (3, 5):  # Matek columns
            cell.font = font_data
        else:
            cell.font = font_data
    row += 1

# Summary rows
for d in t7d_summary:
    for ci, v in enumerate(d, 1):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.alignment = align_right if ci > 1 else align_left
        cell.border = border_all
        cell.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
        cell.font = Font(name='Arial', size=9, bold=True, color='000000')
    row += 1

# ── Column widths ──
ws.column_dimensions['A'].width = 28
for col in range(2, 15):
    ws.column_dimensions[get_column_letter(col)].width = 15

# ── Freeze panes ──
ws.freeze_panes = 'A1'

# ── Print setup ──
ws.page_setup.orientation = 'landscape'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

out_path = r'D:\Navid\Study\Ph.D 2\Thesis\13-self-architecting\asann_article\ASANN_benchmark_table.xlsx'
wb.save(out_path)
print(f'Saved: {out_path}')
print(f'Total rows: {row}')
